import pandas as pd
import math
from io import BytesIO
import requests
import re
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.formatting.rule import FormulaRule
import numpy as np
from constants import Col, Status, FundName, FundCode, RESULT_COLUMNS, SheetName

def normalize_names(df: pd.DataFrame) -> pd.DataFrame:
    """
    이름 열의 공백을 제거하는 공통 헬퍼 함수.
    """
    if Col.NAME in df.columns:
        df[Col.NAME] = df[Col.NAME].astype(str).str.strip()
    return df

def get_google_sheets_title(url: str) -> str | None:
    """
    구글 시트 URL에서 제목을 추출한다.
    """
    try:
        if "/edit" in url:
            view_url = url.replace("/edit", "/view")
        else:
            view_url = url

        response = requests.get(view_url, timeout=10)
        response.raise_for_status()

        title_match = re.search(r"<title>(.*?)</title>", response.text)
        if title_match:
            title = title_match.group(1)
            title = title.replace(" - Google Sheets", "").replace(" - Google スプレッドシート", "")
            return title.strip()

        return None
    except Exception:
        return None

def extract_date_from_filename(filename: str) -> tuple[int, int] | None:
    """
    파일명에서 년월을 추출한다.
    패턴: ※정리본_YY.M.D.xlsx (예: ※정리본_26.2.14.xlsx → (2026, 2))
    """
    match = re.search(r"_(\d{2})\.(\d{1,2})\.\d{1,2}\.xlsx?$", filename)
    if not match:
        return None

    yy = int(match.group(1))
    month = int(match.group(2))

    year = 2000 + yy if yy <= 30 else 1900 + yy

    return (year, month)

def extract_first_payment_month(df: pd.DataFrame) -> pd.DataFrame:
    """
    최초 납부월 명단을 추출한다.
    조건:
      - (코드1==1 & 코드2 1~7) 
      - (코드1==2 & 코드2==1) 
      - (코드1==3 & 코드2==1)
    """
    # 0. 데이터 전처리 (이름 공백 제거)
    df_work = normalize_names(df.copy())
        
    # 1. 조건 필터링
    condition = (
        ((df_work[Col.CODE1] == 1) & (df_work[Col.CODE2].between(1, 7))) |
        ((df_work[Col.CODE1].isin([2, 3])) & (df_work[Col.CODE2] == 1))
    ).fillna(False)
    df_first = df_work[condition].copy()
    
    # 2. 불필요한 열 제거
    exclude_cols = ["구분", "코드3"]
    valid_cols = [c for c in df_first.columns if not str(c).startswith("Unnamed") and c not in exclude_cols]
    df_first = df_first[valid_cols]
    
    # 3. 정렬 및 중복 제거 (가장 과거 날짜 유지)
    df_first = df_first.sort_values([Col.NAME, Col.YEAR, Col.MONTH], ascending=[True, True, True])
    df_first = df_first.drop_duplicates(subset=[Col.NAME], keep="first")
    df_first = df_first.reset_index(drop=True)
    df_first.index = df_first.index + 1
    return df_first

def detect_errors(df: pd.DataFrame) -> pd.DataFrame:
    """
    오류를 검출하여 결과를 생성한다. (가속화 버전)
    """
    # 0. 데이터 전처리 (이름 공백 제거)
    target_df = normalize_names(df.copy())

    # 1. 대상 데이터 필터링
    # 조건: (코드1==1 & 코드2 1~7) OR (코드1==2 & 코드2==1) OR (코드1==3 & 코드2==1)
    condition = (
        ((target_df[Col.CODE1] == 1) & (target_df[Col.CODE2].between(1, 7))) |
        ((target_df[Col.CODE1] == 2) & (target_df[Col.CODE2] == 1)) |
        ((target_df[Col.CODE1] == 3) & (target_df[Col.CODE2] == 1))
    ).fillna(False)
    target_df = target_df[condition].copy()
    
    # 2. 피벗 테이블 생성: (이름, 해당년, 해당월) 별로 각 코드(1,2,3)의 입금 합계 계산
    pivot = target_df.pivot_table(
        index=[Col.NAME, Col.YEAR, Col.MONTH], 
        columns=Col.CODE1, 
        values=Col.RAW_DEPOSIT, 
        aggfunc="sum", 
        fill_value=0
    )
    
    # 필요한 컬럼이 없으면 생성
    for col in [1, 2, 3]:
        if col not in pivot.columns:
            pivot[col] = 0
            
    # 작업 편의를 위해 컬럼 이름 변경
    pivot.columns = ["운영_입금", "협력_입금", "복지_입금"]
    pivot = pivot.reset_index()
    
    # 원본 데이터에서 해당 그룹에 어떤 코드들이 존재하는지 확인 (입금액 0원 등 특이 케이스 대응)
    existence = target_df.groupby([Col.NAME, Col.YEAR, Col.MONTH])[Col.CODE1].apply(set)
    valid_indices = [idx for idx, codes in existence.items() if 1 in codes and (2 in codes or 3 in codes)]
    
    if not valid_indices:
        return pd.DataFrame()

    # pivot 필터링
    pivot_idx = pivot.set_index([Col.NAME, Col.YEAR, Col.MONTH])
    valid_mask = pivot_idx.index.isin(valid_indices)
    work_df = pivot_idx[valid_mask].copy()
    
    results = []
    
    # 벡터화 연산을 위해 컬럼 추출
    op_deposit = work_df["운영_입금"]
    years = work_df.index.get_level_values(Col.YEAR)
    months = work_df.index.get_level_values(Col.MONTH)
    
    # 3. 공통 검사 루프 (협력기금 & 복지기금)
    fund_configs = [
        {
            "name": FundName.COOPERATION,
            "code": FundCode.COOPERATION,
            "deposit": work_df["협력_입금"],
            "has_mask": pd.Series([2 in c for c in existence[work_df.index]], index=work_df.index),
            "type": "coop"
        },
        {
            "name": FundName.WELFARE,
            "code": FundCode.WELFARE,
            "deposit": work_df["복지_입금"],
            "has_mask": pd.Series([3 in c for c in existence[work_df.index]], index=work_df.index),
            "type": "welf"
        }
    ]

    for config in fund_configs:
        # 기금별 기준금액 및 산출법 설정
        if config["type"] == "coop":
            # 협력기금 비율 결정 (2019년 3월 이전 0.3, 이후 0.4)
            ratio_mask = (years <= 2018) | ((years == 2019) & (months <= 3))
            ratios = pd.Series(np.where(ratio_mask, 0.3, 0.4), index=work_df.index)
            
            target_amounts = op_deposit * ratios
            std = target_amounts.round().astype(int)
        else:
            ratios = pd.Series(1.0, index=work_df.index)
            std = op_deposit

        formula_series = "운영기금 총액(" + op_deposit.apply(lambda x: f"{int(x):,}").astype(str) + ") × " + (ratios * 100).astype(int).astype(str) + "%"

        # 부족 검사
        insufficient_mask = ((config["deposit"] < std) & config["has_mask"]).fillna(False)
        if insufficient_mask.any():
            res = work_df[insufficient_mask].copy()
            res[Col.FUND_NAME] = config["name"]
            res[Col.CODE] = config["code"]
            res[Col.STATUS] = Status.INSUFFICIENT
            res[Col.STANDARD] = std[insufficient_mask]
            res[Col.FORMULA] = formula_series[insufficient_mask]
            res[Col.DEPOSIT] = config["deposit"][insufficient_mask]
            res[Col.DIFF] = res[Col.DEPOSIT] - res[Col.STANDARD]
            results.append(res)

        # 초과 검사
        excess_mask = ((config["deposit"] > std) & config["has_mask"]).fillna(False)
        if excess_mask.any():
            res = work_df[excess_mask].copy()
            res[Col.FUND_NAME] = config["name"]
            res[Col.CODE] = config["code"]
            res[Col.STATUS] = Status.EXCESS
            res[Col.STANDARD] = std[excess_mask]
            res[Col.FORMULA] = formula_series[excess_mask]
            res[Col.DEPOSIT] = config["deposit"][excess_mask]
            res[Col.DIFF] = res[Col.DEPOSIT] - res[Col.STANDARD]
            results.append(res)
        
    if not results:
        return pd.DataFrame()
        
    final_df = pd.concat(results).reset_index()
    final_df[Col.REMARKS] = ""
    final_df[Col.PREV_MONTH_AMT] = None
    final_df[Col.NEXT_MONTH_AMT] = None
    
    # 컬럼 정리
    final_df = final_df[RESULT_COLUMNS]
    
    # 코드 컬럼을 문자열로 변환 (pyarrow 호환성)
    final_df[Col.CODE] = final_df[Col.CODE].astype(str)
    
    # 정렬: 이름 -> 해당년 -> 해당월 -> 기금명(코드 순)
    final_df = final_df.sort_values([Col.NAME, Col.YEAR, Col.MONTH, Col.CODE])
    
    final_df = final_df.reset_index(drop=True)
    final_df.index = final_df.index + 1
    
    return final_df

def generate_missed_months(
    df: pd.DataFrame, 
    df_first: pd.DataFrame | None = None, 
    filename: str | None = None,
    graduation_names: list[str] | None = None,
    start_year: int | None = None,
    start_month: int | None = None,
    end_year: int | None = None,
    end_month: int | None = None,
    exemption_map: dict[str, list[str]] | None = None
) -> tuple[pd.DataFrame, int]:
    """
    미납월(누락된 코드)을 생성한다.
    - 코드 1: 11~17 중 하나라도 있으면 납부로 인정. 모두 없으면 미납.
    - 코드 2: 21이 없으면 미납.
    - 코드 3: 31이 없으면 미납.
    - 단, df_first(최초납부월 정보)가 주어지면, 해당 인원의 최초납부월 이전 데이터는 미납에서 제외한다.
    - filename이 주어지면, 기준년월-3개월 이전의 미납에 "과거 미납분"을 비고에 표기한다.
    - graduation_names, end_year, end_month가 주어지면, 졸업생별로 최초납부월~기준년월까지의
      모든 월에 대해 미납을 검출한다 (원본 데이터에 없는 월도 포함).
    """
    # 1. 대상 데이터 필터링 및 "대표 코드(CanonicalCode)" 부여
    # (코드1==1 & 코드2 1~7) -> 1
    # (코드1==2 & 코드2==1)   -> 2
    # (코드1==3 & 코드2==1)   -> 3
    
    # 먼저 전체 데이터 중 유효한 코드 범위를 가진 것만 남김 (또는 계산을 위해 플래그 생성)
    # 복사를 떠서 작업
    target_df = normalize_names(df.copy())
    
    # Canonical Code 매핑
    target_df["canonical"] = 0
    mask_op = ((target_df[Col.CODE1] == 1) & (target_df[Col.CODE2].between(1, 7))).fillna(False)
    mask_coop = ((target_df[Col.CODE1] == 2) & (target_df[Col.CODE2] == 1)).fillna(False)
    mask_welf = ((target_df[Col.CODE1] == 3) & (target_df[Col.CODE2] == 1)).fillna(False)
    
    target_df.loc[mask_op, "canonical"] = 1
    target_df.loc[mask_coop, "canonical"] = 2
    target_df.loc[mask_welf, "canonical"] = 3
    
    # 유효한 canonical 코드를 가진 행만 남김 (이름/년/월 그룹 파악용)
    valid_rows = target_df[target_df["canonical"] > 0].copy()
    
    # [신규] graduation_names, end_year, end_month가 제공되면 전체 기간 생성
    if graduation_names and end_year and end_month and df_first is not None:
        first_map = df_first.set_index(Col.NAME)[[Col.YEAR, Col.MONTH]]
        first_map["serial"] = first_map[Col.YEAR] * 12 + first_map[Col.MONTH]
        first_serial_dict = first_map["serial"].to_dict()
        
        end_serial = end_year * 12 + end_month
        
        analysis_start_serial = None
        if start_year is not None:
            sm = start_month if start_month else 1
            analysis_start_serial = start_year * 12 + sm
        
        all_periods = []
        for name in graduation_names:
            if name in first_serial_dict:
                start_serial = first_serial_dict[name]
            else:
                name_data = target_df[target_df[Col.NAME] == name]
                if name_data.empty:
                    continue
                start_serial = int(name_data[Col.YEAR].min() * 12 + name_data[Col.MONTH].min())
            
            if analysis_start_serial is not None:
                start_serial = max(start_serial, analysis_start_serial)
            
            for serial in range(start_serial, end_serial + 1):
                year = serial // 12
                month = serial % 12
                if month == 0:
                    year -= 1
                    month = 12
                all_periods.append({Col.NAME: name, Col.YEAR: year, Col.MONTH: month})
        
        if not all_periods:
            return pd.DataFrame(), 0
            
        groups = pd.DataFrame(all_periods)
    else:
        # [기존 로직] 그룹핑 기준: 데이터에 존재하는 (이름, 년, 월)
        groups = target_df[[Col.NAME, Col.YEAR, Col.MONTH]].drop_duplicates()
    
    # 3. 각 그룹별로 필요한 Canonical Code (1, 2, 3)를 Cartesian Product로 생성
    groups["key"] = 1
    codes = pd.DataFrame({"canonical": [1, 2, 3], "key": [1, 1, 1]})
    
    expected = pd.merge(groups, codes, on="key").drop("key", axis=1)
    
    # 4. 실제 존재하는 Canonical Code 확인
    # "납부했다"고 인정받으려면 valid_rows(canonical > 0)에 있어야 함.
    # 따라서 existing은 valid_rows에서 추출
    existing = valid_rows[[Col.NAME, Col.YEAR, Col.MONTH, "canonical"]].drop_duplicates()
    existing["present"] = True
    
    # 5. 병합 및 미납 필터링
    merged = pd.merge(expected, existing, on=[Col.NAME, Col.YEAR, Col.MONTH, "canonical"], how="left")
    missed = merged[merged["present"].isna()].copy()
    
    if missed.empty:
        return pd.DataFrame(), 0

    # [신규] 최초 납부월 이전 데이터 필터링
    filtered_count = 0 
    if df_first is not None and not df_first.empty:
        # 1. df_first에서 (이름) -> (Year * 12 + Month) 매핑 생성
        first_map = df_first.set_index(Col.NAME)[[Col.YEAR, Col.MONTH]]
        first_map["serial"] = first_map[Col.YEAR] * 12 + first_map[Col.MONTH]
        first_serial_dict = first_map["serial"].to_dict()
        
        # 2. missed에 serial 컬럼 추가
        missed["serial"] = missed[Col.YEAR] * 12 + missed[Col.MONTH]
        
        # 3. 필터링 함수
        def is_after_first(row):
            try:
                name = row[Col.NAME]
                if name not in first_serial_dict:
                    return True # 최초 납부 정보 없으면 유지
                return row["serial"] >= first_serial_dict[name]
            except Exception as e:
                # 에러 발생 시 데이터 컨텍스트 포함
                raise ValueError(f"데이터 검증 중 오류 발생 (이름: {row.get(Col.NAME,'?')}, 년월: {row.get(Col.YEAR,'?')}-{row.get(Col.MONTH,'?')}): {e}")
        
        # 마스크 계산
        # [수정] .fillna 시 object 타입 배열의 다운캐스팅 FutureWarning 해결
        mask_series = missed.apply(is_after_first, axis=1)
        # 결과가 object 타입인지 확인 후 결측치를 채우고 bool 타입으로 변환
        mask = mask_series.infer_objects(copy=False).fillna(False).astype(bool)
        # 제외된 건수 계산
        filtered_count = (~mask).sum()
        
        missed = missed[mask]
        
        if missed.empty:
            return pd.DataFrame(), filtered_count
    
    # 면제기금 필터링
    if exemption_map:
        canonical_to_fund = {1: FundName.OPERATING, 2: FundName.COOPERATION, 3: FundName.WELFARE}
        
        def is_exempted(row):
            name = row[Col.NAME]
            canonical = row["canonical"]
            if name not in exemption_map:
                return False
            fund_name = canonical_to_fund.get(canonical, "")
            return fund_name in exemption_map[name]
        
        exempt_mask = missed.apply(is_exempted, axis=1)
        missed = missed[~exempt_mask].copy()
        
        if missed.empty:
            return pd.DataFrame(), filtered_count

    # 파일명 기준년월과 동일한 미납 건 제외
    ref_date = extract_date_from_filename(filename) if filename else None
    if ref_date:
        same_mask = (missed[Col.YEAR] == ref_date[0]) & (missed[Col.MONTH] == ref_date[1])
        missed = missed[~same_mask]
        if missed.empty:
            return pd.DataFrame(), filtered_count

    # 6. 결과 포맷팅
    label_map = {1: FundName.OPERATING, 2: FundName.COOPERATION, 3: FundName.WELFARE}
    code_map = {1: FundCode.OPERATING, 2: FundCode.COOPERATION, 3: FundCode.WELFARE}
    
    missed[Col.FUND_NAME] = missed["canonical"].map(label_map)
    # [수정] 호환되지 않는 dtype 설정 시 발생하는 FutureWarning 해결
    # Col.CODE가 처음부터 object/문자열로 처리되도록 함
    missed[Col.CODE] = missed["canonical"].map(code_map).astype(str)
    
    # [수정] 운영기금(1) 미납 시 코드를 "11~17"로 변경
    missed.loc[missed["canonical"] == 1, Col.CODE] = "11~17"

    missed[Col.DEPOSIT] = 0
    missed[Col.STATUS] = Status.UNPAID

    op_monthly = valid_rows[valid_rows["canonical"] == 1].groupby(
        [Col.NAME, Col.YEAR, Col.MONTH]
    )[Col.RAW_DEPOSIT].sum()

    def _calc_standard(row):
        key = (row[Col.NAME], row[Col.YEAR], row[Col.MONTH])
        canonical = row["canonical"]
        if canonical == 1:
            return pd.NA, ""
        op_total = op_monthly.get(key)
        if op_total is None or op_total == 0:
            return pd.NA, ""
        ratio_mask_val = (row[Col.YEAR] <= 2018) or ((row[Col.YEAR] == 2019) and (row[Col.MONTH] <= 3))
        ratio = 0.3 if ratio_mask_val else 0.4
        std = int(round(op_total * ratio))
        formula = f"운영기금 총액({int(op_total):,}) × {int(ratio * 100)}%"
        return std, formula

    std_formula = missed.apply(_calc_standard, axis=1, result_type="expand")
    missed[Col.STANDARD] = std_formula[0]
    missed[Col.FORMULA] = std_formula[1]
    missed[Col.DIFF] = pd.NA
    
    if ref_date:
        cutoff_serial = ref_date[0] * 12 + ref_date[1] - 2
        if "serial" not in missed.columns:
            missed["serial"] = missed[Col.YEAR] * 12 + missed[Col.MONTH]
        missed[Col.REMARKS] = missed.apply(
            lambda row: "3개월 이전 미납분" if row["serial"] < cutoff_serial else "",
            axis=1
        )
        missed = missed.drop(columns=["serial"])
    else:
        missed[Col.REMARKS] = ""
    missed[Col.PREV_MONTH_AMT] = None
    missed[Col.NEXT_MONTH_AMT] = None
    
    missed = missed[RESULT_COLUMNS].sort_values([Col.NAME, Col.YEAR, Col.MONTH, Col.CODE])
    
    return missed, filtered_count

def to_excel_bytes(df_first, df_errors, output_filename, df_summary=None, df_raw=None):
    """
    데이터프레임을 엑셀 바이트로 변환 (리팩토링 버전)
    """
    buffer = BytesIO()
    
    def _calc_adjacent_month_amounts(df_errors, df_raw):
        if df_raw is None or df_errors is None or df_errors.empty or df_raw.empty:
            return df_errors
        if Col.CODE1 not in df_raw.columns or Col.CODE2 not in df_raw.columns:
            return df_errors
        df_work = df_raw.copy()
        df_work["canonical"] = 0
        mask_op = ((df_work[Col.CODE1] == 1) & (df_work[Col.CODE2].between(1, 7))).fillna(False)
        mask_coop = ((df_work[Col.CODE1] == 2) & (df_work[Col.CODE2] == 1)).fillna(False)
        mask_welf = ((df_work[Col.CODE1] == 3) & (df_work[Col.CODE2] == 1)).fillna(False)
        df_work.loc[mask_op, "canonical"] = 1
        df_work.loc[mask_coop, "canonical"] = 2
        df_work.loc[mask_welf, "canonical"] = 3
        valid = df_work[df_work["canonical"] > 0].copy()
        valid["serial"] = valid[Col.YEAR] * 12 + valid[Col.MONTH]
        grouped = valid.groupby([Col.NAME, "serial", "canonical"])[Col.RAW_DEPOSIT].sum().reset_index()
        lookup = {}
        for _, row in grouped.iterrows():
            lookup[(row[Col.NAME], row["serial"], row["canonical"])] = int(row[Col.RAW_DEPOSIT])
        df_out = df_errors.copy()
        if Col.PREV_MONTH_AMT not in df_out.columns:
            df_out[Col.PREV_MONTH_AMT] = None
        if Col.NEXT_MONTH_AMT not in df_out.columns:
            df_out[Col.NEXT_MONTH_AMT] = None
        for idx, row in df_out.iterrows():
            status = str(row.get(Col.STATUS, ""))
            if status not in (Status.INSUFFICIENT, Status.EXCESS, Status.UNPAID):
                continue
            fund_name = str(row.get(Col.FUND_NAME, ""))
            canonical_map = {"운영": 1, "협력": 2, "복지": 3}
            canonical = canonical_map.get(fund_name)
            if canonical is None:
                continue
            name = row[Col.NAME]
            cur_serial = int(row[Col.YEAR]) * 12 + int(row[Col.MONTH])
            prev_serial = cur_serial - 1
            next_serial = cur_serial + 1
            prev_val = lookup.get((name, prev_serial, canonical))
            next_val = lookup.get((name, next_serial, canonical))
            df_out.at[idx, Col.PREV_MONTH_AMT] = prev_val if prev_val is not None else 0
            df_out.at[idx, Col.NEXT_MONTH_AMT] = next_val if next_val is not None else 0
        return df_out
    
    df_errors = _calc_adjacent_month_amounts(df_errors, df_raw)
    
    def prepare_sheet(df, exclude_cols=None):
        if df is None or df.empty:
            return None
        # Unnamed 컬럼 제거
        cols = [c for c in df.columns if not str(c).startswith("Unnamed")]
        if exclude_cols:
            cols = [c for c in cols if c not in exclude_cols]
        return df[cols]

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:  # type: ignore[arg-type]
        thin_border = Border(
            left=Side(style='thin', color='FF9CA3AF'),
            right=Side(style='thin', color='FF9CA3AF'),
            top=Side(style='thin', color='FF9CA3AF'),
            bottom=Side(style='thin', color='FF9CA3AF')
        )
        thick_bottom_border = Border(
            left=Side(style='thin', color='FF9CA3AF'),
            right=Side(style='thin', color='FF9CA3AF'),
            top=Side(style='thin', color='FF9CA3AF'),
            bottom=Side(style='thick', color='FF9CA3AF')
        )

        # 1. 오류검출결과 시트 (사용자 요청으로 앞에 위치)
        df_e = prepare_sheet(df_errors)
        if df_e is not None:
            df_e.index = range(1, len(df_e) + 1)
            df_e.to_excel(writer, index=True, index_label="번호", sheet_name=SheetName.ERROR_RESULT)
            ws = writer.sheets[SheetName.ERROR_RESULT]
            
            # 컬럼 너비 설정 (비율 조정: 80/110 ≈ 0.727)
            # A: 번호 (80px), B: 이름 (100px), C-E: 기본 (90px)
            # F: 코드 (80px), G,I,K: 납부금액/기준금액/차액 (110px)
            # A: 번호, B: 이름, C-D: 납부년월, E: 기금명, F: 코드
            # G: 상태, H: 차액, I-J: 납부/기준금액, K: 산출법
            # L-M: 전/익월, N: 비고
            ws.column_dimensions["A"].width = 8   # 번호
            ws.column_dimensions["B"].width = 10  # 이름
            ws.column_dimensions["C"].width = 9   # 납부년
            ws.column_dimensions["D"].width = 9   # 납부월
            ws.column_dimensions["E"].width = 9   # 기금명
            ws.column_dimensions["F"].width = 8   # 코드
            ws.column_dimensions["G"].width = 12  # 상태
            ws.column_dimensions["H"].width = 12  # 차액
            ws.column_dimensions["I"].width = 12  # 납부금액
            ws.column_dimensions["J"].width = 12  # 기준금액
            ws.column_dimensions["K"].width = 26  # 기준금액 산출법
            ws.column_dimensions["L"].width = 14  # 전월납부금액
            ws.column_dimensions["M"].width = 14  # 익월납부금액
            ws.column_dimensions["N"].width = 41  # 비고
            
            # 오토필터 및 정렬/포맷팅
            last_col = get_column_letter(len(df_e.columns) + 1)
            ws.auto_filter.ref = f"A1:{last_col}{len(df_e) + 1}"

            header_font = Font(color="FFFFFFFF", bold=True, size=11)
            header_fill = PatternFill(start_color="FF272F3A", end_color="FF272F3A", fill_type="solid")
            header_align = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 30

            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thick_bottom_border

            fills = {}

            year_month_fill = PatternFill(start_color="FFD4E3F8", end_color="FFD4E3F8", fill_type="solid")
            adjacent_month_fill = PatternFill(start_color="FFE0F2F1", end_color="FFE0F2F1", fill_type="solid")
            fund_fills = {
                "협력": PatternFill(start_color="FFFEC28E", end_color="FFFEC28E", fill_type="solid"),
                "복지": PatternFill(start_color="FFB1DCA6", end_color="FFB1DCA6", fill_type="solid"),
                "운영": PatternFill(start_color="FFA8C8E8", end_color="FFA8C8E8", fill_type="solid"),
            }
            even_row_fill = PatternFill(start_color="FFF5F3F4", end_color="FFF5F3F4", fill_type="solid")
            default_font = Font(color="FF000000", size=10)

            # 헤더에 배경색 및 두꺼운 하단 테두리 적용
            for cell in ws[1]:
                col_idx_zero = cell.col_idx - 1
                if col_idx_zero in fills:
                    cell.fill = fills[col_idx_zero]
                cell.border = thick_bottom_border

            last_data_row = len(df_e) + 1

            ws.conditional_formatting.add(
                f"C2:D{last_data_row}",
                FormulaRule(formula=['TRUE'], fill=year_month_fill)
            )
            ws.conditional_formatting.add(
                f"L2:M{last_data_row}",
                FormulaRule(formula=['TRUE'], fill=adjacent_month_fill)
            )
            for fund_name, fund_fill in fund_fills.items():
                ws.conditional_formatting.add(
                    f"E2:E{last_data_row}",
                    FormulaRule(formula=[f'$E2="{fund_name}"'], fill=fund_fill)
                )
            for alt_range in [f"A2:B{last_data_row}", f"F2:K{last_data_row}", f"N2:N{last_data_row}"]:
                ws.conditional_formatting.add(
                    alt_range,
                    FormulaRule(formula=['MOD(SUBTOTAL(3,$A$2:$A2),2)=0'], fill=even_row_fill)
                )

            for row in ws.iter_rows(min_row=2, max_row=last_data_row):
                excel_row_num = row[0].row

                for idx, cell in enumerate(row):
                    cell.font = default_font
                    cell.border = thin_border

                    if idx == 5:
                        pass

                    if idx == 6:
                        status_value = str(cell.value).strip() if cell.value else ""
                        if status_value == "미납":
                            cell.font = Font(color='FF9CA3AF', bold=True)
                            cell.value = "✕ 미납"
                        elif status_value == "부족":
                            cell.font = Font(color='FFEF4444', bold=True)
                            cell.value = "▼ 부족"
                        elif status_value == "초과":
                            cell.font = Font(color='FF059669', bold=True)
                            cell.value = "▲ 초과"

                    if idx == 7:
                        status_cell = ws.cell(row=excel_row_num, column=7)
                        status_value = str(status_cell.value).strip() if status_cell.value else ""
                        if "미납" in status_value:
                            cell.font = Font(color='FF9CA3AF', bold=True)
                        elif "부족" in status_value:
                            cell.font = Font(color='FFEF4444', bold=True)
                        elif "초과" in status_value:
                            cell.font = Font(color='FF059669', bold=True)
                        if cell.value is not None and isinstance(cell.value, (int, float)):
                            cell.number_format = '+#,##0;-#,##0;0'

                    if idx in [8, 9]:
                        cell.alignment = Alignment(horizontal="right")
                        if cell.value is not None:
                            cell.number_format = "#,##0"

                    if idx in [11, 12]:
                        cell.alignment = Alignment(horizontal="right")
                        if cell.value is not None and isinstance(cell.value, (int, float)):
                            cell.number_format = "#,##0"

                    if idx <= 5:
                        cell.alignment = Alignment(horizontal="center")
                    elif idx == 6:
                        cell.alignment = Alignment(horizontal="center")
                    elif idx == 7:
                        cell.alignment = Alignment(horizontal="right")
                    elif idx in [8, 9]:
                        cell.alignment = Alignment(horizontal="right")
                    elif idx == 10:
                        cell.alignment = Alignment(horizontal="left")
                    else:
                        cell.alignment = Alignment(horizontal="center")
            ws.freeze_panes = "A2"

        # 2. 오류요약 시트 (이름별 요약)
        df_s = prepare_sheet(df_summary)
        if df_s is not None:
            df_s.index = range(1, len(df_s) + 1)
            df_s.to_excel(writer, index=True, index_label="번호", sheet_name=SheetName.ERROR_SUMMARY)
            ws_s = writer.sheets[SheetName.ERROR_SUMMARY]
            
            ws_s.row_dimensions[1].height = 30
            for cell in ws_s[1]:
                cell.font = Font(color="FFFFFFFF", bold=True, size=11)
                cell.fill = PatternFill(start_color="FF272F3A", end_color="FF272F3A", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thick_bottom_border
            
            ws_s.column_dimensions["A"].width = 8
            ws_s.column_dimensions["B"].width = 12
            ws_s.column_dimensions["C"].width = 9
            ws_s.column_dimensions["D"].width = 9
            ws_s.column_dimensions["E"].width = 9
            ws_s.column_dimensions["F"].width = 12
            
            last_col = get_column_letter(len(df_s.columns) + 1)
            ws_s.auto_filter.ref = f"A1:{last_col}{len(df_s) + 1}"
            
            for row in ws_s.iter_rows(min_row=2, max_row=len(df_s) + 1):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = thin_border
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "#,##0"
            
            if not df_s.empty:
                subtotal_row = len(df_s) + 2
                ws_s.cell(row=subtotal_row, column=1, value="소계")
                ws_s.cell(row=subtotal_row, column=2, value=len(df_s))
                ws_s.cell(row=subtotal_row, column=3, value=int(df_s["미납"].sum()))
                ws_s.cell(row=subtotal_row, column=4, value=int(df_s["부족"].sum()))
                ws_s.cell(row=subtotal_row, column=5, value=int(df_s["초과"].sum()))
                ws_s.cell(row=subtotal_row, column=6, value=int(df_s["합계"].sum()))
                for cell in ws_s[subtotal_row]:
                    cell.alignment = Alignment(horizontal="center")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "#,##0"
                
                subtotal_values = [cell.value for cell in ws_s[subtotal_row]]
                ws_s.delete_rows(subtotal_row, 1)
                ws_s.insert_rows(2, 1)
                for col_idx, value in enumerate(subtotal_values, start=1):
                    cell = ws_s.cell(row=2, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal="center")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    if isinstance(value, (int, float)):
                        cell.number_format = "#,##0"
                
                ws_s.auto_filter.ref = f"A1:{last_col}{len(df_s) + 2}"
            
            ws_s.freeze_panes = "A3"


        # 3. 최초납부월 시트
        df_f = prepare_sheet(df_first, exclude_cols=["구분", "코드3"])
        if df_f is not None:
            df_f.index = range(1, len(df_f) + 1)
            df_f.to_excel(writer, index=True, index_label="번호", sheet_name=SheetName.FIRST_PAYMENT)
            ws_f = writer.sheets[SheetName.FIRST_PAYMENT]

            ws_f.row_dimensions[1].height = 30
            for cell in ws_f[1]:
                cell.font = Font(color="FFFFFFFF", bold=True, size=11)
                cell.fill = PatternFill(start_color="FF272F3A", end_color="FF272F3A", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thick_bottom_border

            col_names = [str(c) for c in df_f.columns]
            for i, name in enumerate(col_names):
                col_letter = get_column_letter(i + 2)
                if name == "작업일자":
                    ws_f.column_dimensions[col_letter].width = 15
                elif name == "이름":
                    ws_f.column_dimensions[col_letter].width = 10
                elif name == "입금":
                    ws_f.column_dimensions[col_letter].width = 10
                else:
                    ws_f.column_dimensions[col_letter].width = 9
            ws_f.column_dimensions["A"].width = 8

            center_cols = {"납부년", "납부월", "코드1", "코드2", "이름", "작업일자"}
            for row in ws_f.iter_rows(min_row=2, max_row=len(df_f) + 1):
                for cell in row:
                    cell.border = thin_border
                    if cell.column == 1:
                        cell.alignment = Alignment(horizontal="center")
                    elif col_names[cell.column - 2] in center_cols:
                        cell.alignment = Alignment(horizontal="center")
                    if col_names[cell.column - 2] == "입금" and isinstance(cell.value, (int, float)):
                        cell.number_format = "#,##0"

            last_col = get_column_letter(len(df_f.columns) + 1)
            ws_f.auto_filter.ref = f"A1:{last_col}{len(df_f) + 1}"
            ws_f.freeze_panes = "A2"
    buffer.seek(0)
    return buffer.getvalue(), output_filename
